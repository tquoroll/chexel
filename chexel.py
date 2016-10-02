# Import a wrapper for the Companies House API
import chwrapper
# Import a handler for Excel
import openpyxl
# Import a tool to compare two strings for similarity
from difflib import SequenceMatcher
# Import os to pull the API key from environment variable
import os

COMPANIES_HOUSE_KEY = os.environ['COMPANIES_HOUSE_KEY']

# Open the input file and append company names to a list
def source_companies():
    companies_list = []
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb.active

    for col in ws.iter_cols(min_row=2, max_col=1):
        for cell in col:
            companies_list.append(cell.value)

    return companies_list

# Get the list of companies and search each of them in turn
def run_search():

    imported_companies = source_companies()

    print("Imported these companies: %s" % imported_companies)

    chsearch = chwrapper.Search()

    results_list = []

# Search each company in turn and handle Companies House rate limiting
    for company in imported_companies:
        try:
            company_search = chsearch.search_companies(company)
        except HTTPError as e:
            if e.code == 429:
                print("Waiting for the API...")
                time.sleep(300)
                company_search = chsearch.search_companies(company)

# Pick up the JSON file including the company's results
        Company_JSON = (company_search.json())

        co_name = Company_JSON['items'][0]['title']
        co_address = Company_JSON['items'][0]['address_snippet']
        co_number = Company_JSON['items'][0]['company_number']

# Check the similarity of the input name and the result name. Now we just report, in future action if not similar.
        co_confidence = int(SequenceMatcher(None, (company.lower()), co_name.lower()).ratio()*100)

        results_list.append({"Searched": company, "Name": co_name, "Address": co_address, "Number": co_number, "Confidence": co_confidence})

# Print names to show progress.
        if co_confidence > 90:
            print("Processed %s, a good match." % company)
        else:
            print("Processed %s, a poor match." % company)

    return results_list

def excel_out():

    results_list = run_search()

# Create a new workbook to record results.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Search Results"

#Set up headers.
    ws['A1'].value = "Input Name"
    ws['B1'].value = "Company Name"
    ws['C1'].value = "Company Number"
    ws['D1'].value = "Address"
    ws['E1'].value = "Confidence"

    ws['A1'].font = openpyxl.styles.Font(bold=True)
    ws['B1'].font = openpyxl.styles.Font(bold=True)
    ws['C1'].font = openpyxl.styles.Font(bold=True)
    ws['D1'].font = openpyxl.styles.Font(bold=True)
    ws['E1'].font = openpyxl.styles.Font(bold=True)

# Populate spreadsheet with company data.

    r = 2
    for index, company in enumerate(results_list):
        ws.cell(row=r,column=1).value = results_list[index]["Searched"]
        ws.cell(row=r, column=2).value = results_list[index]["Name"]
        ws.cell(row=r, column=3).value = results_list[index]["Number"]
        ws.cell(row=r, column=4).value = results_list[index]["Address"]
        ws.cell(row=r, column=5).value = str(results_list[index]["Confidence"])
        r += 1

# Resize columns to roughly fit contents (best solution I could find...)

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    wb.save(filename='output.xlsx')

excel_out()






